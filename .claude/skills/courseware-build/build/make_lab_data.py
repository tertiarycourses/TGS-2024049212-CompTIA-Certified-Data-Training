#!/usr/bin/env python3
"""Generate the mock datasets shipped with each lab (CSV + XLSX).

Every dataset is written into its lab folder as `data/<name>.csv`, and the labs that
would realistically be opened in a spreadsheet also get an `.xlsx` workbook. Generation
is DETERMINISTIC (fixed seed, no wall-clock) so the numbers quoted in the slides, the
Learner Guide and the assessment model answers always match what the learner sees.

The datasets are deliberately larger than a hand-typed sample: n is big enough that
z-scores, t-tests and correlations behave the way the textbook says they do, which the
original 7-to-10-row samples did not.

Run:  python3 make_lab_data.py
"""
import os, sys, csv, random
from datetime import date, timedelta

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import course_data as C
from data_domain1 import DOMAIN1
from data_domain2 import DOMAIN2
from data_domain3 import DOMAIN3
from data_domain4 import DOMAIN4
from data_domain5 import DOMAIN5
ACT = DOMAIN1 + DOMAIN2 + DOMAIN3 + DOMAIN4 + DOMAIN5


def _find_repo(start):
    env = os.environ.get("COURSE_REPO")
    if env and os.path.isdir(env):
        return env
    d = start
    for _ in range(8):
        d = os.path.dirname(d)
        if os.path.isdir(os.path.join(d, "courseware")) and os.path.isdir(os.path.join(d, "labs")):
            return d
    return os.path.dirname(os.path.dirname(HERE))


REPO = _find_repo(HERE)
LABS = os.path.join(REPO, "labs")

RNG = random.Random(20260819)          # fixed seed — reproducible datasets

FIRST = ["Mei", "Ravi", "Siti", "John", "Wei Ming", "Priya", "Daniel", "Nurul", "Kai",
         "Aishah", "Marcus", "Divya", "Jun Hao", "Farah", "Terence", "Lakshmi", "Ryan",
         "Hui Ling", "Arjun", "Grace", "Zhi Wei", "Sofia", "Ben", "Anitha", "Yee Ling",
         "Hakim", "Chloe", "Vikram", "Serene", "Amirah"]
LAST = ["Tan", "Kumar", "Nur", "Lee", "Lim", "Sharma", "Ong", "Binte Rahman", "Wong",
        "Ismail", "Chan", "Menon", "Goh", "Abdullah", "Ee", "Pillai", "Ng", "Teo",
        "Singh", "Koh", "Chua", "Reyes", "Tay", "Raj", "Foo", "Osman", "Yeo", "Nair",
        "Sim", "Hassan"]
REGIONS = ["Central", "East", "North", "West"]
CITIES = {"Central": "Singapore", "East": "Tampines", "North": "Woodlands", "West": "Jurong"}
DEPTS = ["Operations", "Sales", "Technology", "Finance", "Marketing"]
PRODUCTS = [
    (10, "Wireless Mouse", 24.90), (11, "USB-C Hub", 59.00), (12, "Mechanical Keyboard", 129.00),
    (13, "27in Monitor", 349.00), (14, "Laptop Stand", 45.50), (15, "Webcam 1080p", 78.00),
    (16, "Noise-Cancelling Headset", 219.00), (17, "External SSD 1TB", 165.00),
    (18, "Docking Station", 289.00), (19, "Ergonomic Chair", 459.00),
]


def write_csv(lab_dir, name, header, rows):
    d = os.path.join(lab_dir, "data")
    os.makedirs(d, exist_ok=True)
    p = os.path.join(d, name)
    with open(p, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    return p


def write_xlsx(lab_dir, name, sheets):
    """sheets: list of (sheet_name, header, rows). Header row is bold + frozen."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("   (openpyxl missing — skipping", name, ")")
        return None
    d = os.path.join(lab_dir, "data")
    os.makedirs(d, exist_ok=True)
    p = os.path.join(d, name)
    wb = Workbook()
    wb.remove(wb.active)
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F6FEB")
    for sname, header, rows in sheets:
        ws = wb.create_sheet(sname[:31])
        ws.append(list(header))
        for c in ws[1]:
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        for r in rows:
            ws.append(list(r))
        for i, h in enumerate(header, 1):
            width = max(len(str(h)) + 4,
                        max((len(str(r[i - 1])) for r in rows[:200]), default=8) + 3)
            ws.column_dimensions[get_column_letter(i)].width = min(width, 42)
        ws.freeze_panes = "A2"
    wb.save(p)
    return p


def lab_dir_for(num):
    for d in sorted(os.listdir(LABS)):
        if d.startswith(f"lab-{num:02d}-"):
            return os.path.join(LABS, d)
    raise SystemExit(f"lab folder for lab {num} not found under {LABS}")


written = []


def note(paths):
    for p in paths:
        if p:
            written.append(os.path.relpath(p, REPO))


# ───────────────────────────────────────────── LAB 1 — relational schema (3 tables)
d1 = lab_dir_for(1)
cust_rows = []
for i in range(1, 41):
    fn, ln = FIRST[(i - 1) % len(FIRST)], LAST[(i * 7) % len(LAST)]
    region = REGIONS[i % 4]
    joined = date(2025, 1, 1) + timedelta(days=(i * 9) % 330)
    cust_rows.append([i, fn, ln,
                      f"{fn.split()[0].lower()}.{ln.split()[0].lower()}{i}@example.sg",
                      region, joined.isoformat()])
prod_rows = [[pid, nm, price] for pid, nm, price in PRODUCTS]
order_rows = []
oid = 100
for i in range(1, 121):
    cid = RNG.randint(1, 40)
    pid, _, _ = PRODUCTS[RNG.randrange(len(PRODUCTS))]
    qty = RNG.choice([1, 1, 1, 2, 2, 3])
    od = date(2025, 3, 1) + timedelta(days=RNG.randint(0, 89))
    order_rows.append([oid, cid, pid, qty, od.isoformat()])
    oid += 1
note([
    write_csv(d1, "customers.csv", ["customer_id", "first_name", "last_name", "email", "region", "joined"], cust_rows),
    write_csv(d1, "products.csv", ["product_id", "name", "unit_price"], prod_rows),
    write_csv(d1, "orders.csv", ["order_id", "customer_id", "product_id", "qty", "order_date"], order_rows),
    write_xlsx(d1, "sales-schema.xlsx", [
        ("customers", ["customer_id", "first_name", "last_name", "email", "region", "joined"], cust_rows),
        ("products", ["product_id", "name", "unit_price"], prod_rows),
        ("orders", ["order_id", "customer_id", "product_id", "qty", "order_date"], order_rows)]),
])

# ───────────────────────────────────────────── LAB 2 — structured / semi / unstructured
d2 = lab_dir_for(2)
l2 = []
for i in range(1, 13):
    fn, ln = FIRST[i % len(FIRST)], LAST[(i * 3) % len(LAST)]
    region = REGIONS[i % 4]
    l2.append([i, f"{fn} {ln}", CITIES[region], round(RNG.uniform(45, 420), 2)])
note([write_csv(d2, "customers.csv", ["customer_id", "name", "city", "spend"], l2)])
import json as _json
docs = []
for i, (cid, nm, city, spend) in enumerate(l2):
    rec = {"customer_id": cid, "name": nm, "city": city, "spend": spend}
    if i % 3 == 0:
        rec["tags"] = ["vip", "repeat"] if i % 2 == 0 else ["new"]
    if i % 4 == 0:
        rec["contact"] = {"email": f"cust{cid}@example.sg", "phone": f"9{RNG.randint(1000000, 9999999)}"}
    docs.append(rec)
os.makedirs(os.path.join(d2, "data"), exist_ok=True)
_p = os.path.join(d2, "data", "customers.json")
open(_p, "w").write(_json.dumps(docs, indent=1))
written.append(os.path.relpath(_p, REPO))
_p = os.path.join(d2, "data", "notes.txt")
open(_p, "w").write(
    "Quarterly account notes — unstructured, exactly as the account managers typed them.\n\n"
    + "\n".join(
        f"{nm} from {city} spent about ${spend:.2f} with us this quarter."
        if i % 2 == 0 else
        f"{nm} ({city}) spent {spend:.0f} dollars, and asked about bulk pricing."
        for i, (_, nm, city, spend) in enumerate(l2))
    + "\n")
written.append(os.path.relpath(_p, REPO))

# ───────────────────────────────────────────── LAB 4 — the dirty dataset
d4 = lab_dir_for(4)
dirty = []
for i in range(1, 61):
    fn, ln = FIRST[(i * 5) % len(FIRST)], LAST[(i * 11) % len(LAST)]
    region = REGIONS[i % 4]
    dirty.append([i, f"{fn} {ln}", CITIES[region],
                  f"{RNG.uniform(40, 480):.2f}",
                  (date(2025, 3, 1) + timedelta(days=i % 30)).isoformat()])
# planted defects — the counts below are quoted in the LG and the assessment key
for idx in (7, 22, 41, 53):                      # 4 missing cities
    dirty[idx][2] = ""
for idx in (12, 29, 47):                         # 3 missing spends
    dirty[idx][3] = ""
for idx in (18, 35):                             # 2 missing dates
    dirty[idx][4] = ""
dirty.append(list(dirty[3]))                     # 3 exact duplicate rows
dirty.append(list(dirty[26]))
dirty.append(list(dirty[44]))
dirty[9][3] = "99999.00"                         # 2 extreme outliers
dirty[38][3] = "87500.00"
dirty[15][2] = "  Jurong  "                      # leading/trailing spaces
dirty[31][2] = "singapore"                       # inconsistent casing
note([
    write_csv(d4, "sales_dirty.csv", ["order_id", "customer", "city", "spend", "order_date"], dirty),
    write_xlsx(d4, "sales_dirty.xlsx", [("sales_dirty", ["order_id", "customer", "city", "spend", "order_date"], dirty)]),
])

# ───────────────────────────────────────────── LAB 5 — messy contact strings
d5 = lab_dir_for(5)
contacts = []
for i in range(1, 31):
    fn, ln = FIRST[(i * 2) % len(FIRST)], LAST[(i * 5) % len(LAST)]
    nm = f"{fn} {ln}"
    email = f"{fn.split()[0].lower()}.{ln.split()[0].lower()}@example.sg"
    if i % 3 == 0:
        phone = f"+65 {RNG.choice('689')}{RNG.randint(100,999)} {RNG.randint(1000,9999)}"
    elif i % 3 == 1:
        phone = f"{RNG.choice('689')}{RNG.randint(1000000, 9999999)}"
    else:
        phone = f"+65{RNG.choice('689')}{RNG.randint(1000000, 9999999)}"
    contacts.append([i, f"{nm} <{email}> {phone}"])
note([write_csv(d5, "contacts.csv", ["id", "contact"], contacts)])

# ───────────────────────────────────────────── LAB 6 — CIDR sites
d6 = lab_dir_for(6)
sites = [["HQ", "192.168.10.0/24"], ["Branch-West", "10.0.4.0/22"], ["DMZ", "172.16.8.0/29"],
         ["Warehouse", "192.168.20.0/26"], ["Retail-East", "10.1.0.0/16"],
         ["Lab-Network", "172.20.14.0/28"], ["Guest-WiFi", "192.168.100.0/23"],
         ["Backup-Site", "10.10.10.0/25"], ["IoT-Segment", "172.30.0.0/20"],
         ["Point-to-Point", "192.168.250.0/30"],
         ["Legacy-Import", "10.0.5.0/22"]]   # host bits set on purpose — see the troubleshooting note
note([write_csv(d6, "sites.csv", ["site", "cidr"], sites)])

# ───────────────────────────────────────────── LAB 7 — three sources to integrate
d7 = lab_dir_for(7)
c7 = []
for i in range(1, 31):
    fn, ln = FIRST[(i * 4) % len(FIRST)], LAST[(i * 9) % len(LAST)]
    c7.append([i, f"{fn} {ln}", REGIONS[i % 4]])
c7.append([31, "Farah Osman", "East"])           # a customer with NO orders
o7 = []
for i in range(1, 81):
    cid = RNG.randint(1, 30)
    o7.append([200 + i, cid, f"{RNG.uniform(35, 520):.2f}"])
o7.append([999, 77, "55.00"])                    # an ORPHAN order (customer 77 does not exist)
# targets straddle actual revenue on purpose: Central and West MISS, East and North BEAT,
# so the "% of target" panel in Lab 12 has something real to show.
t7 = [[r, v] for r, v in zip(REGIONS, [6600, 4500, 5400, 5800])]
note([
    write_csv(d7, "customers.csv", ["customer_id", "name", "region"], c7),
    write_csv(d7, "orders.csv", ["order_id", "customer_id", "amount"], o7),
    write_csv(d7, "targets.csv", ["region", "target"], t7),
    write_xlsx(d7, "integration-sources.xlsx", [
        ("customers", ["customer_id", "name", "region"], c7),
        ("orders", ["order_id", "customer_id", "amount"], o7),
        ("targets", ["region", "target"], t7)]),
])

# ───────────────────────────────────────────── LAB 8 — salaries (n=60 + 1 outlier)
d8 = lab_dir_for(8)
BANDS = {"Operations": (3600, 4600), "Sales": (3800, 5200), "Technology": (4600, 6400),
         "Finance": (4200, 5600), "Marketing": (3900, 5100)}
sal = []
for i in range(1, 61):
    dept = DEPTS[i % len(DEPTS)]
    lo, hi = BANDS[dept]
    sal.append([f"E{i:03d}", dept, int(RNG.uniform(lo, hi) // 50 * 50)])
sal.append(["E999", "Executive", 26000])         # the single outlier that moves the mean
note([
    write_csv(d8, "salaries.csv", ["employee", "dept", "salary"], sal),
    write_xlsx(d8, "salaries.xlsx", [("salaries", ["employee", "dept", "salary"], sal)]),
])

# ───────────────────────────────────────────── LAB 9 — A/B test (n=60 per group)
d9 = lab_dir_for(9)
ab = []
for i in range(60):
    ab.append(["A", round(RNG.gauss(40.5, 4.2), 2)])
for i in range(60):
    ab.append(["B", round(RNG.gauss(49.0, 4.6), 2)])
note([
    write_csv(d9, "abtest.csv", ["group", "order_value"], ab),
    write_xlsx(d9, "abtest.xlsx", [("abtest", ["group", "order_value"], ab)]),
])

# ───────────────────────────────────────────── LAB 10 — marketing spend vs revenue
d10 = lab_dir_for(10)
mk = []
for m in range(1, 37):
    spend = 10 + m * 0.85 + RNG.gauss(0, 0.9)
    revenue = 52 + spend * 6.6 + RNG.gauss(0, 7.5)
    staff = 12 + m * 0.28 + RNG.gauss(0, 0.5)
    mk.append([m, round(spend, 1), round(revenue, 1), int(round(staff))])
note([
    write_csv(d10, "marketing.csv", ["month", "spend", "revenue", "staff"], mk),
    write_xlsx(d10, "marketing.xlsx", [("marketing", ["month", "spend", "revenue", "staff"], mk)]),
])

# ───────────────────────────────────────────── LAB 11 & 12 — sales / KPI
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
BASE = {"Central": 118, "West": 88, "North": 142, "East": 96}
TARGET = {"Central": 130, "West": 100, "North": 145, "East": 105}
sales_rows, kpi_rows = [], []
for mi, mo in enumerate(MONTHS):
    for reg in ["Central", "West", "North", "East"]:
        rev = round(BASE[reg] * (1 + 0.055 * mi) + RNG.gauss(0, 4.5), 1)
        orders = int(rev / RNG.uniform(2.6, 3.0))
        unit = round(rev * 1000 / max(orders, 1) / 1000 * 1.0, 2)
        sales_rows.append([mo, reg, rev, orders, unit])
        kpi_rows.append([mo, reg, rev, orders, TARGET[reg]])
d11 = lab_dir_for(11)
note([
    write_csv(d11, "sales.csv", ["month", "region", "revenue", "orders", "unit_price"], sales_rows),
    write_xlsx(d11, "sales.xlsx", [("sales", ["month", "region", "revenue", "orders", "unit_price"], sales_rows)]),
])
d12 = lab_dir_for(12)
note([
    write_csv(d12, "kpi.csv", ["month", "region", "revenue", "orders", "target"], kpi_rows),
    write_xlsx(d12, "kpi.xlsx", [("kpi", ["month", "region", "revenue", "orders", "target"], kpi_rows)]),
])

# ───────────────────────────────────────────── LAB 13 — sensitive PII extract
d13 = lab_dir_for(13)
POSTAL = ["738099", "600123", "310045", "529536", "018956", "159836", "099253",
          "228095", "760501", "520201", "469001", "129588"]
pii = []
for i in range(1, 41):
    fn, ln = FIRST[(i * 6) % len(FIRST)], LAST[(i * 13) % len(LAST)]
    dept = DEPTS[i % len(DEPTS)]
    lo, hi = BANDS[dept]
    pii.append([i, f"{fn} {ln}",
                f"S{RNG.randint(1000000, 9999999)}{RNG.choice('ABCDEFGHIZJ')}",
                f"{fn.split()[0].lower()}.{ln.split()[0].lower()}@example.sg",
                POSTAL[i % len(POSTAL)], dept, int(RNG.uniform(lo, hi) // 50 * 50)])
note([
    write_csv(d13, "customers.csv", ["cust_id", "name", "nric", "email", "postal", "dept", "salary"], pii),
    write_xlsx(d13, "customers.xlsx", [("customers", ["cust_id", "name", "nric", "email", "postal", "dept", "salary"], pii)]),
])

# ───────────────────────────────────────────── LAB 15 — daily feeds (clean + broken)
d15 = lab_dir_for(15)
good = []
for i in range(1, 41):
    good.append([i, 100 + i, (date(2025, 3, 1) + timedelta(days=i % 14)).isoformat(),
                 f"{RNG.uniform(40, 460):.2f}",
                 RNG.choice(["SHIPPED", "SHIPPED", "SHIPPED", "PENDING", "CANCELLED"])])
bad = []
for i in range(41, 81):
    bad.append([i, 100 + i, (date(2025, 3, 15) + timedelta(days=i % 14)).isoformat(),
                f"{RNG.uniform(40, 460):.2f}",
                RNG.choice(["SHIPPED", "PENDING", "CANCELLED"])])
bad[5][1] = ""                                   # completeness  — null customer_id
bad[19][1] = ""
bad[11][0] = bad[10][0]                          # uniqueness    — duplicate order_id
bad[27][0] = bad[26][0]
bad[8][3] = "-45.00"                             # validity      — negative amount
bad[33][3] = "-12.75"
bad[14][4] = "TELEPORTED"                        # consistency   — invalid status
bad[30][4] = "shipped"
bad[22][2] = "15/03/2025"                        # accuracy      — unparseable date
note([
    write_csv(d15, "feed_good.csv", ["order_id", "customer_id", "order_date", "amount", "status"], good),
    write_csv(d15, "feed_bad.csv", ["order_id", "customer_id", "order_date", "amount", "status"], bad),
    write_xlsx(d15, "daily-feeds.xlsx", [
        ("feed_good", ["order_id", "customer_id", "order_date", "amount", "status"], good),
        ("feed_bad", ["order_id", "customer_id", "order_date", "amount", "status"], bad)]),
])

print(f"Wrote {len(written)} data file(s):")
for w in sorted(written):
    print("  ", w)
